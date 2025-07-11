import numpy as np
from PIL import Image
from Metric_torch import *
from natsort import natsorted
from tqdm import tqdm
import os
import torch
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")   # 忽略所有警告
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")    # 选择GPU或CPU


def write_excel(excel_name='metric.xlsx', worksheet_name='VIF', column_index=0, data=None):
    try:
        workbook = load_workbook(excel_name)    # 尝试加载现有Excel文件
    except FileNotFoundError:
        workbook = Workbook()    # 文件不存在则创建新工作簿

    # 获取或创建工作表
    worksheet = workbook.create_sheet(title=worksheet_name) if worksheet_name not in workbook.sheetnames else workbook[
        worksheet_name]

    column = get_column_letter(column_index + 1)    # 获取列字母（A,B,C...）
    for i, value in enumerate(data):
        cell = worksheet[column + str(i + 1)]   # 定位单元格
        cell.value = value      # 写入数据

    # 保存文件
    workbook.save(excel_name)


def evaluation_one(ir_name, vi_name, f_name):
    # 打开并转换图像为灰度图
    f_img = Image.open(f_name).convert('L')
    ir_img = Image.open(ir_name).convert('L')
    vi_img = Image.open(vi_name).convert('L')

    # 转换为PyTorch张量并移到指定设备
    f_img_tensor = torch.tensor(np.array(f_img)).float().to(device)
    ir_img_tensor = torch.tensor(np.array(ir_img)).float().to(device)
    vi_img_tensor = torch.tensor(np.array(vi_img)).float().to(device)

    # 创建不同数据类型的图像数组（用于不同指标计算）
    f_img_int = np.array(f_img).astype(np.int32)
    f_img_double = np.array(f_img).astype(np.float32)

    ir_img_int = np.array(ir_img).astype(np.int32)
    ir_img_double = np.array(ir_img).astype(np.float32)

    vi_img_int = np.array(vi_img).astype(np.int32)
    vi_img_double = np.array(vi_img).astype(np.float32)

    # 计算21种不同的图像融合评估指标
    CE = CE_function(ir_img_tensor, vi_img_tensor, f_img_tensor)
    NMI = NMI_function(ir_img_int, vi_img_int, f_img_int, gray_level=256)
    QNCIE = QNCIE_function(ir_img_tensor, vi_img_tensor, f_img_tensor)
    TE = TE_function(ir_img_tensor, vi_img_tensor, f_img_tensor)
    EI = EI_function(f_img_tensor)
    Qy = Qy_function(ir_img_tensor, vi_img_tensor, f_img_tensor)
    Qcb = Qcb_function(ir_img_tensor, vi_img_tensor, f_img_tensor)
    EN = EN_function(f_img_tensor)
    MI = MI_function(ir_img_int, vi_img_int, f_img_int, gray_level=256)
    SF = SF_function(f_img_tensor)
    SD = SD_function(f_img_tensor)
    AG = AG_function(f_img_tensor)
    PSNR = PSNR_function(ir_img_tensor, vi_img_tensor, f_img_tensor)
    MSE = MSE_function(ir_img_tensor, vi_img_tensor, f_img_tensor)
    VIF = VIF_function(ir_img_tensor, vi_img_tensor, f_img_tensor)
    CC = CC_function(ir_img_tensor, vi_img_tensor, f_img_tensor)
    SCD = SCD_function(ir_img_tensor, vi_img_tensor, f_img_tensor)
    Qabf = Qabf_function(ir_img_double, vi_img_double, f_img_double)
    Nabf = Nabf_function(ir_img_tensor, vi_img_tensor, f_img_tensor)
    SSIM = SSIM_function(ir_img_double, vi_img_double, f_img_double)
    MS_SSIM = MS_SSIM_function(ir_img_double, vi_img_double, f_img_double)

    return CE, NMI, QNCIE, TE, EI, Qy, Qcb, EN, MI, SF, AG, SD, CC, SCD, VIF, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM


if __name__ == '__main__':
    if __name__ == '__main__':
        with_mean = True    # 是否计算平均值
        config = {
            'dataroot': 'data/',  # Change to your local infrared and visible images path # 数据集根目录
            'results_root': 'runs/tardal-ct',  # Change to your local fusion images path    # 融合结果目录
            'dataset': 'm3fd',  # Specify the dataset name  # 数据集名称
            'save_dir': 'Metric/eval'  # Directory for saving metrics    # 评估结果保存目录
        }

        # 构建路径
        ir_dir = os.path.join(config['dataroot'], config['dataset'], 'ir')  # Infrared images directory
        vi_dir = os.path.join(config['dataroot'], config['dataset'], 'vi')  # Visible images directory
        ## f_dir路径为 runs/tardal-ct/m3fd（results_root + dataset）
        # f_dir = os.path.join(config['results_root'], config['dataset'])  # Fusion images directory
        f_dir = os.path.join(config['results_root'])  # Fusion images directory

        print(ir_dir)
        print(vi_dir)
        print(f_dir)

        # 创建保存目录
        os.makedirs(config['save_dir'], exist_ok=True)

        # 获取并排序红外图像文件列表（最多300个）
        filelist = natsorted(os.listdir(ir_dir))[:300]
        # 构建评估结果文件名
        metric_save_name = os.path.join(config['save_dir'], f'metric_{config["dataset"]}.xlsx')  # Metrics file name

        # Change to the directory name of the fusion images you want to evaluate
        # Method_list = [
        #     'BDLFusion', 'CAF', 'CDDFuse', 'CoCoNet', 'DATFuse', 'DDcGAN', 'DDFM',
        #     'DeFusion', 'Densefuse', 'DIDFuse', 'EMMA', 'FusinDN', 'GANMcC',
        #     'IF-FILM', 'IGNet', 'IRFS', 'LRRNet', 'MetaFusion', 'MFEIF', 'MRFS',
        #     'PAIF', 'PMGI', 'PSFusion', 'ReCoNet', 'RFN-Nest', 'SDCFusion',
        #     'SDNet', 'SeAFusion', 'SegMif', 'SHIP', 'SuperFusion', 'SwinFusion',
        #     'TarDAL', 'Text-IF', 'TGFuse', 'TIMFusion', 'U2Fusion', 'UMFusion',
        #     'YDTR', 'FusionGAN', 'DetFusion', 'MoE-Fusion', 'PromptF'
        # ]

        #  # 定义要评估的方法列表（当前只评估TarDAL）
        Method_list = [
            'TarDAL'
        ]
        # Starting index for the method 'BDLFusion'
        start_index = Method_list.index('TarDAL')

    # 遍历所有方法
    for i, Method in enumerate(Method_list[start_index:], start=start_index):
        # 初始化所有指标列表
        CE_list = []
        NMI_list = []
        QNCIE_list = []
        TE_list = []
        EI_list = []
        Qy_list = []
        Qcb_list = []
        EN_list = []
        MI_list = []
        SF_list = []
        AG_list = []
        SD_list = []
        CC_list = []
        SCD_list = []
        VIF_list = []
        MSE_list = []
        PSNR_list = []
        Qabf_list = []
        Nabf_list = []
        SSIM_list = []
        MS_SSIM_list = []
        filename_list = ['']        # 文件名列表（第一行为空）
        # # 当前方法的融合图像目录（runs/tardal-ct\TarDAL）
        # sub_f_dir = os.path.join(f_dir, Method)

        # 我的项目sub_f_dir 为runs/tardal-ct
        sub_f_dir = os.path.join(f_dir)
        print("当前方法的融合图像目录：" +sub_f_dir)

        # 创建进度条
        eval_bar = tqdm(filelist)
        for _, item in enumerate(eval_bar):
            # 构建完整文件路径
            ir_name = os.path.join(ir_dir, item)
            vi_name = os.path.join(vi_dir, item)
            f_name = os.path.join(sub_f_dir, item)

            # 检查融合图像是否存在
            if os.path.exists(f_name):
                print(ir_name, vi_name, f_name)
                # 计算指标并添加到列表
                CE, NMI, QNCIE, TE, EI, Qy, Qcb, EN, MI, SF, AG, SD, CC, SCD, VIF, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM = evaluation_one(ir_name, vi_name, f_name)
                CE_list.append(CE)
                NMI_list.append(NMI)
                QNCIE_list.append(QNCIE)
                TE_list.append(TE)
                EI_list.append(EI)
                Qy_list.append(Qy)
                Qcb_list.append(Qcb)
                EN_list.append(EN)
                MI_list.append(MI)
                SF_list.append(SF)
                AG_list.append(AG)
                SD_list.append(SD)
                CC_list.append(CC)
                SCD_list.append(SCD)
                VIF_list.append(VIF)
                MSE_list.append(MSE)
                PSNR_list.append(PSNR)
                Qabf_list.append(Qabf)
                Nabf_list.append(Nabf)
                SSIM_list.append(SSIM)
                MS_SSIM_list.append(MS_SSIM)
                filename_list.append(item)  # 添加文件名
                # 更新进度条
                eval_bar.set_description("{} | {}".format(Method, item))

        # 计算平均值（如果需要）
        if with_mean:
            CE_tensor = torch.tensor(CE_list).mean().item()
            CE_list.append(CE_tensor)
            NMI_tensor = torch.tensor(NMI_list).mean().item()
            NMI_list.append(NMI_tensor)
            QNCIE_tensor = torch.tensor(QNCIE_list).mean().item()
            QNCIE_list.append(QNCIE_tensor)
            TE_tensor = torch.tensor(TE_list).mean().item()
            TE_list.append(TE_tensor)
            EI_tensor = torch.tensor(EI_list).mean().item()
            EI_list.append(EI_tensor)
            Qy_tensor = torch.tensor(Qy_list).mean().item()
            Qy_list.append(Qy_tensor)
            Qcb_tensor = torch.tensor(Qcb_list).mean().item()
            Qcb_list.append(Qcb_tensor)
            EN_tensor = torch.tensor(EN_list).mean().item()
            EN_list.append(EN_tensor)
            MI_tensor = torch.tensor(MI_list).mean().item()
            MI_list.append(MI_tensor)
            SF_tensor = torch.tensor(SF_list).mean().item()
            SF_list.append(SF_tensor)
            AG_tensor = torch.tensor(AG_list).mean().item()
            AG_list.append(AG_tensor)
            SD_tensor = torch.tensor(SD_list).mean().item()
            SD_list.append(SD_tensor)
            CC_tensor = torch.tensor(CC_list).mean().item()
            CC_list.append(CC_tensor)
            SCD_tensor = torch.tensor(SCD_list).mean().item()
            SCD_list.append(SCD_tensor)
            VIF_tensor = torch.tensor(VIF_list).mean().item()
            VIF_list.append(VIF_tensor)
            MSE_tensor = torch.tensor(MSE_list).mean().item()
            MSE_list.append(MSE_tensor)
            PSNR_tensor = torch.tensor(PSNR_list).mean().item()
            PSNR_list.append(PSNR_tensor)
            Qabf_list.append(np.mean(Qabf_list))
            Nabf_tensor = torch.tensor(Nabf_list).mean().item()
            Nabf_list.append(Nabf_tensor)
            SSIM_tensor = torch.tensor(SSIM_list).mean().item()
            SSIM_list.append(SSIM_tensor)
            MS_SSIM_tensor = torch.tensor(MS_SSIM_list).mean().item()
            MS_SSIM_list.append(MS_SSIM_tensor)
            # 添加"mean"标识
            filename_list.append('mean')

        # 在每列开头添加方法名称
        CE_list.insert(0, '{}'.format(Method))
        NMI_list.insert(0, '{}'.format(Method))
        QNCIE_list.insert(0, '{}'.format(Method))
        TE_list.insert(0, '{}'.format(Method))
        EI_list.insert(0, '{}'.format(Method))
        Qy_list.insert(0, '{}'.format(Method))
        Qcb_list.insert(0, '{}'.format(Method))
        EN_list.insert(0, '{}'.format(Method))
        MI_list.insert(0, '{}'.format(Method))
        SF_list.insert(0, '{}'.format(Method))
        AG_list.insert(0, '{}'.format(Method))
        SD_list.insert(0, '{}'.format(Method))
        CC_list.insert(0, '{}'.format(Method))
        SCD_list.insert(0, '{}'.format(Method))
        VIF_list.insert(0, '{}'.format(Method))
        MSE_list.insert(0, '{}'.format(Method))
        PSNR_list.insert(0, '{}'.format(Method))
        Qabf_list.insert(0, '{}'.format(Method))
        Nabf_list.insert(0, '{}'.format(Method))
        SSIM_list.insert(0, '{}'.format(Method))
        MS_SSIM_list.insert(0, '{}'.format(Method))

        # 如果是第一个方法，创建Excel表头
        if i == start_index:
            write_excel(metric_save_name, 'CE', 0, filename_list)
            write_excel(metric_save_name, 'NMI', 0, filename_list)
            write_excel(metric_save_name, 'QNCIE', 0, filename_list)
            write_excel(metric_save_name, 'TE', 0, filename_list)
            write_excel(metric_save_name, 'EI', 0, filename_list)
            write_excel(metric_save_name, 'Qy', 0, filename_list)
            write_excel(metric_save_name, 'Qcb', 0, filename_list)
            write_excel(metric_save_name, 'EN', 0, filename_list)
            write_excel(metric_save_name, "MI", 0, filename_list)
            write_excel(metric_save_name, "SF", 0, filename_list)
            write_excel(metric_save_name, "AG", 0, filename_list)
            write_excel(metric_save_name, "SD", 0, filename_list)
            write_excel(metric_save_name, "CC", 0, filename_list)
            write_excel(metric_save_name, "SCD", 0, filename_list)
            write_excel(metric_save_name, "VIF", 0, filename_list)
            write_excel(metric_save_name, "MSE", 0, filename_list)
            write_excel(metric_save_name, "PSNR", 0, filename_list)
            write_excel(metric_save_name, "Qabf", 0, filename_list)
            write_excel(metric_save_name, "Nabf", 0, filename_list)
            write_excel(metric_save_name, "SSIM", 0, filename_list)
            write_excel(metric_save_name, "MS_SSIM", 0, filename_list)

        # 将结果写入Excel
        write_excel(metric_save_name, 'CE', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in CE_list])
        write_excel(metric_save_name, 'NMI', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in NMI_list])
        write_excel(metric_save_name, 'QNCIE', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in QNCIE_list])
        write_excel(metric_save_name, 'TE', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in TE_list])
        write_excel(metric_save_name, 'EI', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in EI_list])
        write_excel(metric_save_name, 'Qy', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in Qy_list])
        write_excel(metric_save_name, 'Qcb', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in Qcb_list])
        write_excel(metric_save_name, 'EN', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in EN_list])
        write_excel(metric_save_name, 'MI', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in MI_list])
        write_excel(metric_save_name, 'SF', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in SF_list])
        write_excel(metric_save_name, 'AG', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in AG_list])
        write_excel(metric_save_name, 'SD', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in SD_list])
        write_excel(metric_save_name, 'CC', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in CC_list])
        write_excel(metric_save_name, 'SCD', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in SCD_list])
        write_excel(metric_save_name, 'VIF', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in VIF_list])
        write_excel(metric_save_name, 'MSE', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in MSE_list])
        write_excel(metric_save_name, 'PSNR', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in PSNR_list])
        write_excel(metric_save_name, 'Qabf', i + 1, Qabf_list)
        write_excel(metric_save_name, 'Nabf', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in Nabf_list])
        write_excel(metric_save_name, 'SSIM', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in SSIM_list])
        write_excel(metric_save_name, 'MS_SSIM', i + 1,
                    [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                     in MS_SSIM_list])
